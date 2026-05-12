"""Local sync server for the wolf-pelt data table.

Eliminates the manual download → drop → update.bat cycle. Accepts POSTs from
`data_table.html` (admin mode) and writes directly into the project folder.
Runs `step1d_dataqc.py` in the background after each save and exposes the
result to the UI for live feedback.

Usage:
    python sync_server.py     # serves on http://127.0.0.1:7869

Security: bound to localhost only. The browser page detects whether the
server is alive on load and switches to live-sync mode if so; otherwise the
page works exactly as before (localStorage + manual download).

Backups: every xlsx / data_decisions.json write creates a timestamped copy
under `.sync_backups/`. Last 20 per filename are kept.
"""

from __future__ import annotations

import json
import os
import shutil
import subprocess
import sys
import threading
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path

PROJECT_DIR = Path(__file__).parent.resolve()
DATA_DECISIONS_PATH = PROJECT_DIR / "data_decisions.json"
XLSX_PATH = PROJECT_DIR / "wolves_data.xlsx"
SHEET_NAME = "נתוני זיהוי זאבים (2)"
BACKUP_DIR = PROJECT_DIR / ".sync_backups"
LOG_PATH = PROJECT_DIR / "sync_server.log"
PIPELINE_SCRIPT = PROJECT_DIR / "step1d_dataqc.py"
PORT = 7869
HOST = "127.0.0.1"

# Concurrency guards
_xlsx_lock = threading.Lock()
_decisions_lock = threading.Lock()
_pipeline_lock = threading.Lock()

# Last pipeline status (read by /api/status)
_last_pipeline = {
    "running": False,
    "ok": None,
    "returncode": None,
    "stdout_tail": "",
    "stderr_tail": "",
    "started_at": None,
    "ended_at": None,
}


# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

def log(msg: str) -> None:
    ts = datetime.now().isoformat(timespec="seconds")
    line = f"[{ts}] {msg}\n"
    try:
        with open(LOG_PATH, "a", encoding="utf-8") as f:
            f.write(line)
    except Exception:
        pass
    print(line, end="", flush=True)


# ---------------------------------------------------------------------------
# Safe file IO
# ---------------------------------------------------------------------------

def atomic_write(path: Path, content: bytes) -> None:
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_bytes(content)
    os.replace(tmp, path)


def backup_file(path: Path) -> None:
    if not path.exists():
        return
    BACKUP_DIR.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = BACKUP_DIR / f"{path.stem}.{ts}{path.suffix}"
    shutil.copy2(path, backup)
    backups = sorted(BACKUP_DIR.glob(f"{path.stem}.*{path.suffix}"))
    if len(backups) > 20:
        for b in backups[:-20]:
            try:
                b.unlink()
            except Exception:
                pass


# ---------------------------------------------------------------------------
# Save handlers
# ---------------------------------------------------------------------------

def save_decisions(payload: dict) -> dict:
    """Merge incoming decisions into data_decisions.json (incoming wins)."""
    with _decisions_lock:
        if DATA_DECISIONS_PATH.exists():
            try:
                existing = json.loads(DATA_DECISIONS_PATH.read_text(encoding="utf-8"))
            except json.JSONDecodeError:
                existing = {"generated_at": "", "schema_version": 1, "decisions": {}}
        else:
            existing = {"generated_at": "", "schema_version": 1, "decisions": {}}
        if not isinstance(existing.get("decisions"), dict):
            existing["decisions"] = {}
        incoming = payload.get("decisions") or {}
        # incoming wins for each id
        existing["decisions"].update(incoming)
        existing["generated_at"] = datetime.now().isoformat() + "Z"
        existing["schema_version"] = 1
        backup_file(DATA_DECISIONS_PATH)
        atomic_write(
            DATA_DECISIONS_PATH,
            json.dumps(existing, ensure_ascii=False, indent=2).encode("utf-8"),
        )
        log(f"save_decisions: merged {len(incoming)} entries, total {len(existing['decisions'])}")
        return {"ok": True, "merged": len(incoming), "total": len(existing["decisions"])}


def save_cell(payload: dict) -> dict:
    """Update a single cell in wolves_data.xlsx and save."""
    import openpyxl
    row_index = payload["row_index"]  # 0-based DataFrame index
    column = payload["column"]
    value = payload.get("value", "")
    with _xlsx_lock:
        backup_file(XLSX_PATH)
        wb = openpyxl.load_workbook(XLSX_PATH)
        if SHEET_NAME not in wb.sheetnames:
            return {"ok": False, "error": f"sheet '{SHEET_NAME}' not found"}
        ws = wb[SHEET_NAME]
        # Build header → column index map
        header = {}
        for c in range(1, ws.max_column + 1):
            h = ws.cell(row=1, column=c).value
            if h is not None:
                header[str(h).strip()] = c
        if column not in header:
            return {"ok": False, "error": f"column '{column}' not found in sheet"}
        col_idx = header[column]
        # row_index is 0-based; xlsx data starts at row 2 (row 1 is header)
        xlsx_row = int(row_index) + 2
        cell_value = None if value == "" else value
        ws.cell(row=xlsx_row, column=col_idx, value=cell_value)
        wb.save(XLSX_PATH)
        log(f"save_cell: row {row_index} col {column} = {value!r}")
        return {"ok": True}


def save_xlsx_bulk(payload: dict) -> dict:
    """Replace sheet (2) wholesale from a JSON snapshot of rows + columns."""
    import openpyxl
    rows = payload.get("rows", [])
    columns = payload.get("columns", [])
    if not columns:
        return {"ok": False, "error": "columns missing"}
    with _xlsx_lock:
        backup_file(XLSX_PATH)
        wb = openpyxl.load_workbook(XLSX_PATH)
        # Rewrite the sheet
        if SHEET_NAME in wb.sheetnames:
            idx = wb.sheetnames.index(SHEET_NAME)
            del wb[SHEET_NAME]
        else:
            idx = 0
        ws = wb.create_sheet(SHEET_NAME, idx)
        for c, name in enumerate(columns, start=1):
            ws.cell(row=1, column=c, value=name)
        for r, row in enumerate(rows, start=2):
            for c, name in enumerate(columns, start=1):
                v = row.get(name)
                if v == "" or v is None:
                    continue
                ws.cell(row=r, column=c, value=v)
        wb.save(XLSX_PATH)
        log(f"save_xlsx_bulk: wrote {len(rows)} rows × {len(columns)} cols")
        return {"ok": True, "rows": len(rows)}


# ---------------------------------------------------------------------------
# Pipeline runner (background)
# ---------------------------------------------------------------------------

def run_pipeline_async() -> None:
    def _run():
        with _pipeline_lock:
            _last_pipeline.update({
                "running": True,
                "started_at": datetime.now().isoformat(),
                "ended_at": None,
                "ok": None,
                "returncode": None,
                "stdout_tail": "",
                "stderr_tail": "",
            })
            try:
                p = subprocess.run(
                    [sys.executable, "-X", "utf8", str(PIPELINE_SCRIPT)],
                    cwd=str(PROJECT_DIR),
                    capture_output=True,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                    timeout=120,
                )
                _last_pipeline["returncode"] = p.returncode
                _last_pipeline["stdout_tail"] = (p.stdout or "")[-800:]
                _last_pipeline["stderr_tail"] = (p.stderr or "")[-800:]
                _last_pipeline["ok"] = (p.returncode == 0)
                if p.returncode == 0:
                    log("pipeline: step1d_dataqc OK")
                else:
                    log(f"pipeline: step1d_dataqc FAILED rc={p.returncode}")
            except subprocess.TimeoutExpired:
                _last_pipeline["ok"] = False
                _last_pipeline["stderr_tail"] = "TIMEOUT (>120s)"
                log("pipeline: TIMEOUT")
            except Exception as e:
                _last_pipeline["ok"] = False
                _last_pipeline["stderr_tail"] = str(e)
                log(f"pipeline: EXCEPTION {e}")
            finally:
                _last_pipeline["running"] = False
                _last_pipeline["ended_at"] = datetime.now().isoformat()

    threading.Thread(target=_run, daemon=True).start()


def _count_decisions() -> int:
    if not DATA_DECISIONS_PATH.exists():
        return 0
    try:
        data = json.loads(DATA_DECISIONS_PATH.read_text(encoding="utf-8"))
        return len(data.get("decisions") or {})
    except Exception:
        return 0


# ---------------------------------------------------------------------------
# HTTP handler
# ---------------------------------------------------------------------------

class SyncHandler(BaseHTTPRequestHandler):
    def _cors(self) -> None:
        # Allow file:// (which sends Origin: null) and any localhost origin.
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")

    def _json(self, status: int, body: dict) -> None:
        b = json.dumps(body, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(b)))
        self._cors()
        self.end_headers()
        try:
            self.wfile.write(b)
        except (BrokenPipeError, ConnectionResetError):
            pass

    def do_OPTIONS(self) -> None:
        self.send_response(204)
        self._cors()
        self.end_headers()

    def do_GET(self) -> None:
        if self.path == "/api/ping":
            self._json(200, {
                "ok": True,
                "version": "1.0",
                "project_dir": str(PROJECT_DIR),
            })
        elif self.path == "/api/status":
            self._json(200, {
                "ok": True,
                "pipeline": dict(_last_pipeline),
                "decisions_count": _count_decisions(),
                "xlsx_mtime": XLSX_PATH.stat().st_mtime if XLSX_PATH.exists() else None,
            })
        else:
            self._json(404, {"ok": False, "error": f"unknown GET {self.path}"})

    def do_POST(self) -> None:
        length = int(self.headers.get("Content-Length", 0))
        raw = self.rfile.read(length) if length else b"{}"
        try:
            payload = json.loads(raw.decode("utf-8"))
        except json.JSONDecodeError as e:
            self._json(400, {"ok": False, "error": f"bad JSON: {e}"})
            return
        try:
            if self.path == "/api/save_decisions":
                result = save_decisions(payload)
                run_pipeline_async()
                self._json(200, result)
            elif self.path == "/api/save_cell":
                result = save_cell(payload)
                run_pipeline_async()
                self._json(200, result)
            elif self.path == "/api/save_xlsx_bulk":
                result = save_xlsx_bulk(payload)
                run_pipeline_async()
                self._json(200, result)
            elif self.path == "/api/run_pipeline":
                run_pipeline_async()
                self._json(200, {"ok": True, "started": True})
            else:
                self._json(404, {"ok": False, "error": f"unknown POST {self.path}"})
        except Exception as e:
            log(f"ERROR {self.path}: {type(e).__name__}: {e}")
            self._json(500, {"ok": False, "error": str(e)})

    def log_message(self, format, *args):
        # Suppress default per-request logging; our log() handles it.
        return


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    BACKUP_DIR.mkdir(exist_ok=True)
    log(f"Sync server starting on http://{HOST}:{PORT}")
    log(f"Project dir: {PROJECT_DIR}")
    log(f"Backups   : {BACKUP_DIR}")
    log("Endpoints : GET /api/ping, GET /api/status, "
        "POST /api/save_decisions, POST /api/save_cell, "
        "POST /api/save_xlsx_bulk, POST /api/run_pipeline")
    server = ThreadingHTTPServer((HOST, PORT), SyncHandler)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        log("Shutting down (Ctrl+C)")
        server.shutdown()


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
